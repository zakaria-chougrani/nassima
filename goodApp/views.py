from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
import json

from openpyxl.worksheet.table import Table, TableStyleInfo

from goodApp.models import Parent, Enfant
from openpyxl import Workbook


# Create your views here.
@login_required(login_url='login')
def home(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        nomComplet = data.get('nomComplet')
        cnie = data.get('cnie')
        etatFamilial = data.get('etatFamilial')
        tailleMenage = data.get('tailleMenage')
        habitat = data.get('habitat')
        adresse = data.get('adresse')
        tel = data.get('tel')
        enfants = data.get('enfants')

        parent = Parent(nomComplet=nomComplet, cnie=cnie, etatFamilial=etatFamilial, tailleMenage=tailleMenage,
                        habitat=habitat, adresse=adresse, tel=tel)
        parent.save()

        for enfant in enfants:
            enfant = Enfant(parent=parent, genre=enfant.get('genre'), age=enfant.get('age'),
                            etablissement=enfant.get('etablissement'), nScol=enfant.get('nScol'))
            enfant.save()
        response_data = {'message': 'Données reçues et enregistrées avec succès.'}
        return JsonResponse(response_data)
    else:
        return render(request, 'home.html')


@login_required(login_url='login')
def dashboard_page(request):
    nbr_mere = Parent.objects.count()
    nbr_enfant = Enfant.objects.count()
    return render(request, 'dashboard.html', {
        'nbr_mere': nbr_mere,
        'nbr_enfant': nbr_enfant
    })


@login_required(login_url='login')
def filtre_age_enfant_view(request):
    if request.method == 'POST':
        age_debut = request.POST.get('ageDebut')
        age_fin = request.POST.get('ageFin')
        if age_debut and age_fin:
            data_search = Enfant.objects.values('parent__nomComplet', 'parent__cnie', 'parent__tel', 'genre',
                                                'age').filter(
                age__gte=age_debut, age__lte=age_fin).order_by('age')

            return render(request, 'filtre.html', {
                'data_table': data_search,
            })
    data = Enfant.objects.values('parent__nomComplet', 'parent__cnie', 'parent__tel', 'genre', 'age').order_by('age')
    return render(request, 'filtre.html', {
        'data_table': data,
    })


@login_required(login_url='login')
def export_csv(request):
    data = Enfant.objects.select_related('parent').all()
    # Créez un nouveau classeur Excel et ajoutez une feuille
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'ID MERE'
    ws['B1'] = 'NOM MERE'
    ws['C1'] = 'CNIE'
    ws['D1'] = 'ETAT FAMILIAL'
    ws['E1'] = 'TAILLE MENAGE'
    ws['F1'] = 'TYPE HABITAT'
    ws['G1'] = 'ADRESSE'
    ws['H1'] = 'TEL MERE'
    ws['I1'] = 'ID ENFANT'
    ws['J1'] = 'GENRE ENFANT'
    ws['K1'] = 'AGE ENFANT'
    ws['L1'] = 'NIVEAU SCOLAIRE ENFANT'
    ws['M1'] = 'ETABLISSEMENT SCOLAIRE ENFANT'
    # Ajoutez les données à la feuille
    for i, row in enumerate(data):
        print(row)
        ws.cell(row=i + 2, column=1, value=row.parent.id)
        ws.cell(row=i + 2, column=2, value=row.parent.nomComplet)
        ws.cell(row=i + 2, column=3, value=row.parent.cnie)
        ws.cell(row=i + 2, column=4, value=row.parent.etatFamilial)
        ws.cell(row=i + 2, column=5, value=row.parent.tailleMenage)
        ws.cell(row=i + 2, column=6, value=row.parent.habitat)
        ws.cell(row=i + 2, column=7, value=row.parent.adresse)
        ws.cell(row=i + 2, column=8, value=row.parent.tel)
        ws.cell(row=i + 2, column=9, value=row.id)
        ws.cell(row=i + 2, column=10, value=row.genre)
        ws.cell(row=i + 2, column=11, value=row.age)
        ws.cell(row=i + 2, column=12, value=row.nScol)
        ws.cell(row=i + 2, column=13, value=row.etablissement)
    # set column widths
    column_widths = [10, 20, 15, 20, 10, 30, 30, 15, 10, 10, 10, 30, 30]  # set desired column widths
    for i, width in enumerate(column_widths):
        ws.column_dimensions[chr(65 + i)].width = width  # adjust column width based on index
    # format the data as a table
    tab = Table(displayName="Table1", ref="A1:M" + str(len(data) + 1))
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Renvoyez le fichier Excel en réponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'
    wb.save(response)
    return response
